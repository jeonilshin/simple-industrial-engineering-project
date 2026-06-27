"""Export endpoints: predictions → .xlsx, schedule → .pdf."""
from __future__ import annotations

from datetime import date, datetime
from io import BytesIO

from fastapi import APIRouter, Depends
from fastapi.responses import StreamingResponse
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill
from reportlab.lib import colors
from reportlab.lib.pagesizes import LETTER
from reportlab.lib.styles import getSampleStyleSheet
from reportlab.lib.units import inch
from reportlab.platypus import (
    Paragraph,
    SimpleDocTemplate,
    Spacer,
    Table,
    TableStyle,
)
from sqlalchemy.orm import Session

from ..auth import get_current_user
from ..database import get_db
from ..models import Employee, Prediction, Schedule

router = APIRouter(
    prefix="/exports",
    tags=["exports"],
    dependencies=[Depends(get_current_user)],
)


@router.get("/predictions.xlsx")
def export_predictions(db: Session = Depends(get_db)):
    rows = (
        db.query(Prediction)
        .order_by(Prediction.created_at.desc())
        .limit(500)
        .all()
    )

    wb = Workbook()
    ws = wb.active
    ws.title = "Predictions"
    headers = [
        "ID",
        "Created (UTC)",
        "Production Target",
        "Predicted Workers",
        "Predicted Cost",
        "Confidence",
    ]
    ws.append(headers)
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill("solid", fgColor="4F46E5")
    for col_idx in range(1, len(headers) + 1):
        cell = ws.cell(row=1, column=col_idx)
        cell.font = header_font
        cell.fill = header_fill

    for r in rows:
        ws.append(
            [
                r.id,
                r.created_at.strftime("%Y-%m-%d %H:%M"),
                r.production_target,
                r.predicted_workers,
                round(r.predicted_cost, 2),
                f"{r.confidence * 100:.1f}%",
            ]
        )

    for col_idx, _ in enumerate(headers, start=1):
        ws.column_dimensions[chr(64 + col_idx)].width = 22

    buffer = BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    filename = f"predictions-{date.today().isoformat()}.xlsx"
    return StreamingResponse(
        buffer,
        media_type=(
            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
        ),
        headers={"Content-Disposition": f"attachment; filename={filename}"},
    )


@router.get("/schedule.pdf")
def export_schedule(db: Session = Depends(get_db)):
    today = date.today()
    rows = (
        db.query(Schedule, Employee)
        .join(Employee, Employee.id == Schedule.employee_id)
        .filter(Schedule.schedule_date == today)
        .order_by(Schedule.shift, Employee.skill_level.desc())
        .all()
    )

    buffer = BytesIO()
    doc = SimpleDocTemplate(buffer, pagesize=LETTER, title="Workforce Schedule")
    styles = getSampleStyleSheet()
    story = []

    story.append(Paragraph("AI Workforce Schedule", styles["Title"]))
    story.append(
        Paragraph(
            f"Date: {today.isoformat()} · Generated: "
            f"{datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}",
            styles["Normal"],
        )
    )
    story.append(Spacer(1, 0.25 * inch))

    if not rows:
        story.append(
            Paragraph(
                "No schedule has been generated for today yet.",
                styles["Italic"],
            )
        )
    else:
        by_shift: dict[str, list[tuple[Schedule, Employee]]] = {}
        for s, e in rows:
            by_shift.setdefault(s.shift, []).append((s, e))

        for shift in ("morning", "afternoon", "night"):
            entries = by_shift.get(shift, [])
            story.append(
                Paragraph(
                    f"{shift.title()} Shift — {len(entries)} workers",
                    styles["Heading2"],
                )
            )
            data = [["Name", "Skill", "Rate", "Hours"]]
            for s, e in entries:
                data.append(
                    [e.name, str(e.skill_level), f"P{e.hourly_rate:.0f}", str(s.hours)]
                )
            table = Table(data, hAlign="LEFT", colWidths=[2.6 * inch, 0.9 * inch, 1.0 * inch, 0.8 * inch])
            table.setStyle(
                TableStyle(
                    [
                        ("BACKGROUND", (0, 0), (-1, 0), colors.HexColor("#4F46E5")),
                        ("TEXTCOLOR", (0, 0), (-1, 0), colors.white),
                        ("FONTNAME", (0, 0), (-1, 0), "Helvetica-Bold"),
                        ("GRID", (0, 0), (-1, -1), 0.25, colors.grey),
                        ("ROWBACKGROUNDS", (0, 1), (-1, -1), [colors.white, colors.HexColor("#F1F5F9")]),
                        ("VALIGN", (0, 0), (-1, -1), "MIDDLE"),
                        ("LEFTPADDING", (0, 0), (-1, -1), 6),
                    ]
                )
            )
            story.append(table)
            story.append(Spacer(1, 0.2 * inch))

    doc.build(story)
    buffer.seek(0)
    return StreamingResponse(
        buffer,
        media_type="application/pdf",
        headers={
            "Content-Disposition": f"attachment; filename=schedule-{today.isoformat()}.pdf"
        },
    )
